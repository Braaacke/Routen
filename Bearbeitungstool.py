"""Interaktives Routenbearbeitungstool mit Zoom-abhängiger Markersichtbarkeit, TSP-Optimierung, Excel- und PDF-Export (mit OSM-Hintergrund)"""

import streamlit as st
import pandas as pd
import leafmap.foliumap as leafmap
import folium
from folium.plugins import MarkerCluster
import networkx as nx
import osmnx as ox
import pickle
from networkx.algorithms.approximation.traveling_salesman import greedy_tsp
from urllib.parse import quote_plus
from datetime import timedelta
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side
from math import radians, sin, cos, sqrt, asin
import re
import matplotlib.pyplot as plt
import contextily as ctx

# Papierformate (Breite × Höhe in Zoll) und default Zoomstufen für Basemap
FORMAT_SIZES = {
    "A4": (8.27, 11.69),
    "A3": (11.69, 16.54),
    "A2": (16.54, 23.39),
    "A1": (23.39, 33.11),
    "A0": (33.11, 46.81),
}
DEFAULT_ZOOMS = {
    "A4": 16,
    "A3": 16,
    "A2": 15,
    "A1": 14,
    "A0": 13,
}

# Funktion für PDF-Export mit variabler Größe, DPI und Zoom
from shapely.geometry import Point, LineString
import geopandas as gpd

def export_routes_pdf_osm(df_assign, filename="routen_uebersicht.pdf", figsize=(8.27, 11.69), dpi=300, zoom=15):
    import matplotlib.pyplot as plt
    import contextily as ctx
    from shapely.geometry import Point, LineString
    import geopandas as gpd
    # Load OSM graph
    graph = get_graph()
    # Prepare point GeoDataFrame in Web Mercator
    pts_gdf = gpd.GeoDataFrame(
        df_assign,
        geometry=gpd.points_from_xy(df_assign['lon'], df_assign['lat']),
        crs='EPSG:4326'
    ).to_crs(epsg=3857)
    # Prepare line features for routes
    line_features = []
    for t in sorted(df_assign['team'].dropna().astype(int).unique()):
        df_t = df_assign[df_assign['team'] == t]
        if 'tsp_order' in df_t.columns:
            df_t = df_t.sort_values('tsp_order')
        pts = list(zip(df_t['lon'], df_t['lat']))
        if len(pts) > 1:
            # get nearest nodes and path
            coords = []
            nodes = [ox.distance.nearest_nodes(graph, X=lon, Y=lat) for lon, lat in pts]
            for u, v in zip(nodes[:-1], nodes[1:]):
                try:
                    path = nx.shortest_path(graph, u, v, weight='length')
                    coords += [(graph.nodes[n]['x'], graph.nodes[n]['y']) for n in path]
                except:
                    pass
            if coords:
                line_features.append({'team': t, 'geometry': LineString(coords)})
    gdf_lines = gpd.GeoDataFrame(line_features, crs='EPSG:4326').to_crs(epsg=3857)
    # Set up figure and axis
    fig, ax = plt.subplots(figsize=figsize)
    # compute bounds
    if not pts_gdf.empty:
        minx, miny, maxx, maxy = pts_gdf.total_bounds
        dx, dy = maxx - minx, maxy - miny
        buf = 0.05
        ax.set_xlim(minx - dx * buf, maxx + dx * buf)
        ax.set_ylim(miny - dy * buf, maxy + dy * buf)
    # draw high-res OSM tiles
    prov = ctx.providers.OpenStreetMap.Mapnik.copy()
    prov['tile_scale'] = 2
    ctx.add_basemap(ax, crs=pts_gdf.crs.to_string(), source=prov, zoom=zoom)
    # draw routes
    colors = ['magenta','cyan','lime','red','orange','yellow','turquoise','purple','pink','blue',
              'black','green','brown','violet','gold','deepskyblue','indigo','crimson','darkorange','teal']
    for idx, row in gdf_lines.iterrows():
        line = row.geometry
        xs, ys = line.xy
        ax.plot(xs, ys, color=colors[idx % len(colors)], linewidth=2, zorder=5)
        # label at midpoint
        mid = line.interpolate(0.5, normalized=True)
        ax.text(mid.x, mid.y, str(int(row['team'])), fontsize=8, color='black',
                fontweight='bold', ha='center', va='center',
                bbox=dict(boxstyle='round,pad=0.2', fc='white', alpha=0.85, lw=1), zorder=6)
    # draw points
    pts_gdf.plot(ax=ax, color='k', markersize=10, zorder=7)
    # draw district boundaries and labels
    try:
        tags = {'place': ['suburb','neighbourhood','quarter']}
        dist = ox.geometries_from_place('Münster, Germany', tags=tags)
        dist = dist[dist['name'].notna() & dist.geometry.type.isin(['Polygon','MultiPolygon'])]
        dist = dist.to_crs(epsg=3857)
        dist.boundary.plot(ax=ax, linewidth=0.8, edgecolor='gray', zorder=3)
        for _, drow in dist.iterrows():
            pt = drow.geometry.representative_point()
            ax.text(pt.x, pt.y, drow['name'], fontsize=6, color='gray', ha='center', va='center',
                    zorder=4, bbox=dict(boxstyle='round,pad=0.1', fc='white', alpha=0.6, lw=0))
    except:
        pass
    # finalize plot
    ax.set_axis_off()
    ax.set_title('Routenübersicht Kontrollbezirke', fontsize=18, fontweight='bold', pad=15)
    plt.tight_layout()
    plt.savefig(filename, bbox_inches='tight', dpi=dpi)
    plt.close(fig)
    return filename


def haversine(lon1, lat1, lon2, lat2):
    dlon = radians(lon2 - lon1)
    dlat = radians(lat2 - lat1)
    a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
    return 2 * 6371000 * asin(sqrt(a))

@st.cache_resource
def get_graph():
    with open("munster_graph.pickle", "rb") as f:
        return pickle.load(f)

def tsp_solve_route(graph, stops_df):
    if len(stops_df) <= 2:
        return stops_df
    coords = list(zip(stops_df["lat"], stops_df["lon"]))
    nodes = [ox.distance.nearest_nodes(graph, X=lon, Y=lat) for lat, lon in coords]
    G = nx.complete_graph(len(nodes))
    for i in G.nodes:
        for j in G.nodes:
            if i != j:
                try:
                    length = nx.shortest_path_length(graph, nodes[i], nodes[j], weight='length')
                    G[i][j]['weight'] = length
                except:
                    G[i][j]['weight'] = float('inf')
    tsp_path = greedy_tsp(G)
    return stops_df.iloc[tsp_path].reset_index(drop=True)

def make_export(df_assign):
    output = io.BytesIO()
    graph = get_graph()
    overview = []
    sheets = {}
    teams = df_assign['team'].dropna().astype(int).unique()
    centers = {t: (
        df_assign[df_assign['team']==t]['lat'].mean(),
        df_assign[df_assign['team']==t]['lon'].mean()
    ) for t in teams}
    sorted_teams = sorted(centers.keys(), key=lambda t: (-centers[t][0], centers[t][1]))
    for idx, t in enumerate(sorted_teams, start=1):
        df_t = df_assign[df_assign['team'] == t]
        if 'tsp_order' in df_t.columns:
            df_t = df_t.sort_values('tsp_order')
        rooms = df_t.get('num_rooms', pd.Series()).sum()
        travel_km = travel_min = 0.0
        coords = df_t[['lat','lon']].values.tolist()
        if len(coords) > 1:
            for (lat1, lon1), (lat2, lon2) in zip(coords[:-1], coords[1:]):
                try:
                    n1 = ox.distance.nearest_nodes(graph, X=lon1, Y=lat1)
                    n2 = ox.distance.nearest_nodes(graph, X=lon2, Y=lat2)
                    length = nx.shortest_path_length(graph, n1, n2, weight='length')
                except:
                    length = haversine(lon1, lat1, lon2, lat2)
                travel_km += length / 1000.0
                travel_min += (length / 1000.0) * 2.0
        ctrl_time = int(rooms * 10)
        total_time = travel_min + ctrl_time
        overview.append({
            'Kontrollbezirk': idx,
            'Anzahl Wahllokale': len(df_t),
            'Anzahl Stimmbezirke': rooms,
            'Wegstrecke (km)': round(travel_km,1),
            'Fahrtzeit (min)': int(travel_min),
            'Kontrollzeit (min)': ctrl_time,
            'Gesamtzeit': str(timedelta(minutes=int(total_time))),
            'Google-Link': 'https://www.google.com/maps/dir/' + '/'.join([f"{lat},{lon}" for lat,lon in coords])
        })
        detail = []
        for j, (_, r) in enumerate(df_t.iterrows(), start=1):
            rooms_str = r.get('rooms','')
            numbers = re.findall(r'\b(\d+)\b', rooms_str)
            stimmbez = ', '.join(numbers)
            detail.append({
               'Nr.': j,
               'Wahllokal': r.get('Wahlraum-B',''),
               'Adresse': r['Wahlraum-A'],
               'Stimmbezirke': stimmbez,
            })
        sheets[idx] = pd.DataFrame(detail)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(overview).to_excel(writer, sheet_name='Übersicht', index=False)
        ws = writer.book.create_sheet('Gesamtübersicht')
        row_cursor = 1
        num_cols = 4
        data_fill = PatternFill(fill_type='solid', start_color='F2F2F2')
        thick = Side(style='thick')
        thin = Side(style='thin')
        for kb, df_s in sheets.items():
            title_cell = ws.cell(row=row_cursor, column=1, value=f"Kontrollbezirk {kb}")
            title_cell.font = title_cell.font.copy(bold=True)
            title_cell.fill = data_fill
            title_cell.border = Border(top=thick, left=thick, right=thin)
            headers = ['Wahllokal', 'Adresse', 'Stimmbezirk']
            for idx_h, hdr_text in enumerate(headers, start=2):
                hdr = ws.cell(row=row_cursor, column=idx_h, value=hdr_text)
                hdr.font = hdr.font.copy(bold=True)
                hdr.fill = data_fill
                left = thin
                right = thick if idx_h==num_cols else thin
                hdr.border = Border(top=thick, left=left, right=right)
            row_cursor += 1
            for j, r in enumerate(df_s.itertuples(), start=1):
                vals = [j, r.Wahllokal, r.Adresse, r.Stimmbezirke]
                for col_idx, val in enumerate(vals, start=1):
                    c = ws.cell(row=row_cursor, column=col_idx, value=val)
                    c.fill = data_fill
                    left = thick if col_idx==1 else thin
                    right = thick if col_idx==num_cols else thin
                    c.border = Border(left=left, right=right)
                row_cursor += 1
            for col in range(1, num_cols+1):
                c = ws.cell(row=row_cursor-1, column=col)
                left = thick if col==1 else thin
                right = thick if col==num_cols else thin
                c.border = Border(bottom=thick, left=left, right=right)
            row_cursor += 1
        for kb, df_s in sheets.items():
            df_s.to_excel(writer, sheet_name=f"Bezirk_{kb}", index=False)
        for ws in writer.sheets.values():
            for col_cells in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
                col_letter = get_column_letter(col_cells[0].column)
                ws.column_dimensions[col_letter].width = max_length + 2
    output.seek(0)
    return output

st.set_page_config(layout='wide')
if 'show_new' not in st.session_state:
    st.session_state.show_new = False
if 'base_addresses' not in st.session_state:
    base = pd.read_csv('cleaned_addresses.csv').reset_index(drop=True)
    xls = pd.read_excel('routes_optimized.xlsx', sheet_name=None)
    tmp = []
    for name, df in xls.items():
        m = re.match(r"Bezirk_(\d+)", name)
        if m and 'Adresse' in df.columns:
            team_id = int(m.group(1))
            for addr in df['Adresse']:
                tmp.append((addr, team_id))
    assign_df = pd.DataFrame(tmp, columns=['Wahlraum-A','team'])
    merged = base.merge(assign_df, on='Wahlraum-A', how='left')
    st.session_state.base_addresses = merged.copy()
    st.session_state.new_assignments = merged.copy()

df_assign = st.session_state.new_assignments.copy()
if 'latlong' in df_assign.columns and ('lat' not in df_assign.columns or 'lon' not in df_assign.columns):
    df_assign[['lat','lon']] = df_assign['latlong'].str.split(',',expand=True).astype(float)
with st.sidebar:
    st.title('Bearbeitung Kontrollbezirke')
    opts = df_assign.dropna(subset=['Wahlraum-B','Wahlraum-A'])
    addrs_search = opts.apply(lambda r: f"{r['Wahlraum-B']} - {r['Wahlraum-A']}", axis=1).tolist()
    st.selectbox(
        'Wahllokal oder Adresse suchen',
        options=[''] + addrs_search,
        index=0,
        format_func=lambda x: 'Wahllokal oder Adresse suchen' if x=='' else x,
        key='search_selection'
    )
    uploaded = st.file_uploader('Alternative Zuweisung importieren', type=['xlsx'])
    if uploaded:
        # import logic
        imp = pd.read_excel(uploaded, sheet_name=None)
        temp = []
        for sheet, df in imp.items():
            if sheet != 'Übersicht' and 'Adresse' in df.columns:
                team_id = int(sheet.split('_')[1])
                for addr in df['Adresse']:
                    temp.append((addr, team_id))
        assigns = pd.DataFrame(temp, columns=['Wahlraum-A','team'])
        st.session_state.new_assignments = (
            st.session_state.base_addresses.drop(columns=['team'])
            .merge(assigns, on='Wahlraum-A', how='left')
        )
        st.success('Import erfolgreich.')
    opts_assign = st.session_state.new_assignments.dropna(subset=['Wahlraum-B','Wahlraum-A'])
    addrs_assign = opts_assign.apply(lambda r: f"{r['Wahlraum-B']} - {r['Wahlraum-A']}", axis=1).tolist()
    sel = st.multiselect('Wahllokal wählen', options=addrs_assign, placeholder='Auswählen')
    teams = sorted(st.session_state.new_assignments['team'].dropna().astype(int).unique())
    tgt = st.selectbox('Kontrollbezirk wählen', options=[None] + teams, placeholder='Auswählen')
    if st.button('Zuweisung übernehmen') and tgt and sel:
        graph = get_graph()
        for label in sel:
            addr = label.split(' - ',1)[1]
            idx = st.session_state.new_assignments[st.session_state.new_assignments['Wahlraum-A']==addr].index[0]
            st.session_state.new_assignments.at[idx,'team'] = tgt
        df_team = st.session_state.new_assignments[st.session_state.new_assignments['team']==tgt]
        opt = tsp_solve_route(graph, df_team)
        st.session_state.new_assignments.loc[opt.index,'tsp_order'] = range(len(opt))
        st.success('Zuweisung gesetzt.')
    if not st.session_state.show_new:
        if st.button('Neuen Kontrollbezirk erstellen'):
            st.session_state.show_new = True
            st.experimental_rerun()
    else:
        max_t = int(st.session_state.new_assignments['team'].max(skipna=True) or 0) + 1
        with st.form('new_district_form'):
            st.markdown(f"### Neuen Kontrollbezirk {max_t} erstellen")
            sel2 = st.multiselect(f"Stops für Kontrollbezirk {max_t}", options=addrs_assign)
            if st.form_submit_button('Erstellen und zuweisen') and sel2:
                graph = get_graph()
                for label in sel2:
                    addr = label.split(' - ',1)[1]
                    idx = st.session_state.new_assignments[st.session_state.new_assignments['Wahlraum-A']==addr].index[0]
                    st.session_state.new_assignments.at[idx,'team'] = max_t
                df_nt = st.session_state.new_assignments[st.session_state.new_assignments['team']==max_t]
                opt2 = tsp_solve_route(graph, df_nt)
                st.session_state.new_assignments.loc[opt2.index,'tsp_order'] = range(len(opt2))
                st.success(f'Kontrollbezirk {max_t} erstellt.')
                st.session_state.show_new = False
    if st.button('Routen berechnen', key='recalc_routes'):
        graph = get_graph()
        for team_id in sorted(st.session_state.new_assignments['team'].dropna().astype(int).unique()):
            df_team = st.session_state.new_assignments[st.session_state.new_assignments['team']==team_id]
            opt = tsp_solve_route(graph, df_team)
            st.session_state.new_assignments.loc[opt.index,'tsp_order'] = range(len(opt))
        st.success('Routen neu berechnet.')

    # Excel-Export
    export_buf = make_export(st.session_state.new_assignments)
    st.download_button(
        'Kontrollbezirke herunterladen',
        data=export_buf,
        file_name='routen_zuweisung.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # GeoJSON-Export der Routen
    line_features = []
    for t in sorted(st.session_state.new_assignments['team'].dropna().astype(int).unique()):
        df_t = st.session_state.new_assignments[st.session_state.new_assignments['team']==t]
        if 'tsp_order' in df_t.columns:
            df_t = df_t.sort_values('tsp_order')
        pts = df_t[['lat','lon']].values.tolist()
        if len(pts) > 1:
            nodes = [ox.distance.nearest_nodes(get_graph(), X=lon, Y=lat) for lat, lon in pts]
            for u, v in zip(nodes[:-1], nodes[1:]):
                try:
                    path_nodes = nx.shortest_path(get_graph(), u, v, weight='length')
                    coords = [(get_graph().nodes[n]['x'], get_graph().nodes[n]['y']) for n in path_nodes]  # lon, lat
                    line = LineString(coords)
                    line_features.append({'team': t, 'geometry': line})
                except:
                    pass
    gdf_lines = gpd.GeoDataFrame(line_features, crs='EPSG:4326')
    geojson_str = gdf_lines.to_json()
    st.download_button(
        label='GeoJSON herunterladen',
        data=geojson_str,
        file_name='routen.geojson',
        mime='application/geo+json'
    )

    # Karte als A3 PDF (600 dpi) exportieren
    if st.button('Karte als A3 PDF (600dpi) exportieren'):
        with st.spinner('Erstelle A3-PDF-Karte, bitte warten...'):
            pdf_file = export_routes_pdf_osm(
                st.session_state.new_assignments,
                figsize=FORMAT_SIZES['A3'],
                dpi=600,
                zoom=DEFAULT_ZOOMS['A3']
            )
        st.success('A3 PDF-Karte erstellt!')
        with open(pdf_file, 'rb') as f:
            st.download_button(
                label='A3 PDF-Karte herunterladen',
                data=f,
                file_name='routen_karte_A3.pdf',
                mime='application/pdf'
            )

# Funktion zum Zeichnen der interaktiven Karte
def draw_map(df_assign):
    search_sel = st.session_state.get('search_selection','')
    if search_sel:
        addr = search_sel.split(' - ',1)[1]
        row = df_assign[df_assign['Wahlraum-A']==addr]
        center = [row.iloc[0]['lat'], row.iloc[0]['lon']] if not row.empty else [df_assign['lat'].mean(), df_assign['lon'].mean()]
        zoom = 17
    else:
        center = [df_assign['lat'].mean(), df_assign['lon'].mean()]
        zoom = 10
    m = leafmap.Map(center=center, zoom=zoom)
    graph = get_graph()
    colors = ['#FF00FF','#00FFFF','#00FF00','#FF0000','#FFA500','#FFFF00','#00CED1','#DA70D6','#FF69B4','#8A2BE2']
    for i, t in enumerate(sorted(df_assign['team'].dropna().unique())):
        df_t = df_assign[df_assign['team']==t]
        if 'tsp_order' in df_t.columns:
            df_t = df_t.sort_values('tsp_order')
        pts = df_t[['lat','lon']].values.tolist()
        if len(pts)>1:
            path=[]
            nodes=[ox.distance.nearest_nodes(graph,X=lon,Y=lat) for lat,lon in pts]
            for u,v in zip(nodes[:-1],nodes[1:]):
                try:
                    p=nx.shortest_path(graph,u,v,weight='length')
                    path.extend([(graph.nodes[n]['y'],graph.nodes[n]['x']) for n in p])
                except:
                    pass
            folium.PolyLine(path,color=colors[i%len(colors)],weight=6,opacity=0.8,
                            tooltip=f"Kontrollbezirk {int(t)}").add_to(m)
        cluster = MarkerCluster(disableClusteringAtZoom=13)
    for _, r in df_assign.dropna(subset=['lat','lon']).iterrows():
        popup_html = (
            f"<div style='white-space: nowrap;'>"
            f"<b>{r['Wahlraum-B']}</b><br>{r['Wahlraum-A']}<br>Anzahl Räume: {r.get('num_rooms','')}"
            "</div>"
        )
        popup = folium.Popup(popup_html, max_width=300)
        marker = folium.Marker(location=[r['lat'], r['lon']], popup=popup)
        cluster.add_child(marker)
    cluster.add_to(m)
    if not st.session_state.get('search_selection',''):
        m.fit_bounds(df_assign[['lat','lon']].values.tolist())
    m.to_streamlit(use_container_width=True, height=700)

# Karte rendern
# ----------------------
draw_map(df_assign)
